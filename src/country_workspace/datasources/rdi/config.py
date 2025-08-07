from collections.abc import Iterable
from enum import Enum, StrEnum
from io import BytesIO
from typing import NotRequired, TypedDict

from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill
from mimetypes import guess_type

from django.conf import settings

from country_workspace.utils.config import BatchNameConfig, ValidateModeConfig
from country_workspace.utils.fields import Record


RDI = str | BytesIO
Sheet = Iterable[Record]
MultiSheet = Iterable[tuple[int, Sheet]]


mime, __ = guess_type("test.xlsx")

EMAIL_SUBJECT = "RDI file processed with errors"
EMAIL_BODY = "Please find the RDI file processed with errors attached."
EMAIL_FROM = settings.DEFAULT_FROM_EMAIL


class Config(BatchNameConfig, ValidateModeConfig):
    master_detail: bool
    beneficiary_id_column: str
    send_to: str
    household_id_column: NotRequired[str]
    household_label: NotRequired[str]
    people_prefix: NotRequired[str]
    first_line: int


class SheetName(StrEnum):
    HOUSEHOLDS = "Households"
    INDIVIDUALS = "Individuals"
    PEOPLE = "People"


class CellStyle(TypedDict):
    font: Font
    fill: PatternFill


class ErrorConfig:
    ERRORS_COLUMN_NAME = "Errors"
    GENERAL_ERRORS_ROW_NAME = "GENERAL ERRORS:"
    DCT_KEY = "dct"
    GENERAL_KEY = "-"

    class StyleType(Enum):
        ROW_ERROR: CellStyle = {"font": Font(bold=False), "fill": PatternFill(fgColor="FFCCCC", fill_type="solid")}
        GENERAL_ERROR: CellStyle = {"font": Font(bold=True), "fill": PatternFill(fgColor="FFE6CC", fill_type="solid")}

        def apply_to(self, cell: Cell) -> None:
            cell.font = self.value["font"]
            cell.fill = self.value["fill"]

        def apply_to_cells(self, cells: Iterable[Cell]) -> None:
            for cell in cells:
                self.apply_to(cell)
