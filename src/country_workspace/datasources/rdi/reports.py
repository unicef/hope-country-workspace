from contextlib import closing
from io import BytesIO
from typing import Any, Mapping, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from django.core.files.base import ContentFile
from django.core.mail import EmailMessage

from country_workspace.models import Household, Individual
from country_workspace.storages import MEDIA_STORAGE
from .config import SheetName, ErrorConfig, mime, EMAIL_SUBJECT, EMAIL_BODY, EMAIL_FROM


def get_headers(ws: Worksheet) -> list[str | None]:
    (row_values,) = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    return list(row_values or [])


def get_col_num_by_name(ws: Worksheet, col_name: str) -> int:
    headers = get_headers(ws)
    index_map = {h: i for i, h in enumerate(headers, start=1)}
    try:
        return index_map[col_name]
    except KeyError:
        raise ValueError(f"Column {col_name} not found in sheet {ws.title}")


def add_errors_column(worksheet: Worksheet) -> int:
    headers = get_headers(worksheet)
    if ErrorConfig.ERRORS_COLUMN_NAME in headers:
        return headers.index(ErrorConfig.ERRORS_COLUMN_NAME) + 1

    col_num = worksheet.max_column + 1
    cell = worksheet.cell(1, col_num, ErrorConfig.ERRORS_COLUMN_NAME)
    ErrorConfig.StyleType.ROW_ERROR.apply_to(cell)
    return col_num


def add_general_errors_row(worksheet: Worksheet, error_col_num: int, general_errors: Iterable[str]) -> None:
    if not general_errors:
        return
    worksheet.insert_rows(2)
    cell = worksheet.cell(2, error_col_num)
    cell.value = " | ".join(sorted(general_errors, key=str))
    label = worksheet.cell(2, 1, ErrorConfig.GENERAL_ERRORS_ROW_NAME)
    ErrorConfig.StyleType.GENERAL_ERROR.apply_to_cells((cell, label))


def fill_error_cell(cell: Cell, data: dict) -> bool:
    if not data:
        return False
    msgs = []
    for key, values in data.items():
        text = " ; ".join(map(str, values)) if isinstance(values, list) else str(values)
        msgs.append(text if key == ErrorConfig.DCT_KEY else f"{key}: {text}")
    if result := " | ".join(msgs):
        cell.value = result
        ErrorConfig.StyleType.ROW_ERROR.apply_to(cell)
    return bool(result)


def process_sheet_errors(ws: Worksheet, id_col_name: str, errors_dict: dict[int, dict], first_line: int) -> None:
    if not errors_dict:
        return

    id_col = get_col_num_by_name(ws, id_col_name)
    err_col = add_errors_column(ws)
    general_errors = set()

    for (cell,) in ws.iter_rows(min_row=first_line, min_col=id_col, max_col=id_col):
        val = cell.value
        if val is None:
            continue

        try:
            entity_id = int(val)
        except (TypeError, ValueError):
            continue

        raw = errors_dict.get(entity_id)
        if not raw:
            continue

        entity_errors = dict(raw)
        if gen := entity_errors.pop(ErrorConfig.GENERAL_KEY, None):
            general_errors.update(gen if isinstance(gen, list) else [gen])

        fill_error_cell(ws.cell(cell.row, err_col), entity_errors)

    add_general_errors_row(ws, err_col, general_errors)


def collect_household_errors(household_mapping: Mapping[int, Any]) -> dict[int, dict]:
    return {hh_id: household.errors for hh_id, household in household_mapping.items() if household.errors}


def collect_individual_errors(individual_mapping: Mapping[int, Any]) -> dict[int, dict]:
    if not individual_mapping:
        return {}
    pk_to_original = {obj.pk: original_id for original_id, obj in individual_mapping.items()}
    qs = Individual.objects.filter(pk__in=pk_to_original).exclude(errors={})
    return {pk_to_original[pk]: errors for pk, errors in qs.values_list("pk", "errors")}


def save_and_send_errors_file(workbook: Workbook, base_filename: str, send_to: str | None) -> str:
    filename = f"errors_{base_filename}"

    with BytesIO() as out:
        workbook.save(out)
        out.seek(0)
        content = out.read()

    path = MEDIA_STORAGE.save(filename, ContentFile(content))

    if send_to:
        email = EmailMessage(
            subject=EMAIL_SUBJECT,
            body=EMAIL_BODY,
            from_email=EMAIL_FROM,
            to=[send_to],
        )
        email.attach(filename, content, mime)
        email.send(fail_silently=False)

    return path


def generate_errors_report(
    base_filename: str,
    config: dict | None = None,
    *,
    households: Mapping[int, Household] | None = None,
    individuals: Mapping[int, Individual] | None = None,
    people: Mapping[int, Individual] | None = None,
) -> str | None:
    if not config:
        return None
    if not households and not (individuals or people):
        return None

    sources = (
        (households, SheetName.HOUSEHOLDS, collect_household_errors),
        (individuals, SheetName.INDIVIDUALS, collect_individual_errors),
        (people, SheetName.PEOPLE, collect_individual_errors),
    )
    errors = {
        sheet_name: sheet_errors
        for mapping, sheet_name, collector in sources
        if mapping and (sheet_errors := collector(mapping))
    }
    if not errors:
        return None

    try:
        with closing(load_workbook(base_filename)) as workbook:
            first_line = config.get("first_line", 1)

            for sheet_name, sheet_errors in errors.items():
                id_column = (
                    config.get("household_id_column")
                    if sheet_name == SheetName.HOUSEHOLDS
                    else config.get("beneficiary_id_column")
                )
                if id_column:
                    process_sheet_errors(workbook[sheet_name], id_column, sheet_errors, first_line)

            return save_and_send_errors_file(workbook, base_filename, config.get("send_to"))

    except Exception as e:
        raise RuntimeError(f"Failed to generate or deliver the error report: {e}") from e
