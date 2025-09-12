import io
from typing import TYPE_CHECKING, Any
from unittest import mock

import openpyxl
import pytest
import xlsxwriter
from django.urls import reverse
from testutils.factories import FlexFieldFactory
from testutils.utils import select_office
from webtest import Upload

from country_workspace.state import state
from country_workspace.workspaces.admin.cleaners.bulk_update import TYPES, create_bulk_update_template
from tests.workspace.actions import stub

if TYPE_CHECKING:
    from django_webtest import DjangoTestApp
    from django_webtest.pytest_plugin import MixinWithInstanceVariables
    from webtest import Checkbox

    from country_workspace.models import AsyncJob
    from country_workspace.workspaces.models import CountryHousehold

pytestmark = [pytest.mark.admin, pytest.mark.smoke, pytest.mark.django_db]


@pytest.fixture
def app(django_app_factory: "MixinWithInstanceVariables") -> "DjangoTestApp":
    from testutils.factories import SuperUserFactory

    django_app = django_app_factory(csrf_checks=False)
    admin_user = SuperUserFactory(username="superuser")
    django_app.set_user(admin_user)
    django_app._user = admin_user
    return django_app


@pytest.fixture
def mock_media_storage():
    return mock.MagicMock(
        **{
            "save.return_value": "mocked/path/to/file",
            "exists.return_value": False,
            "get_available_name.return_value": "mocked/path/to/file",
            "open.return_value": io.BytesIO(b"mocked file content"),
        }
    )


@pytest.fixture
def office():
    from testutils.factories import OfficeFactory

    co = OfficeFactory()
    state.tenant = co
    return co


@pytest.fixture(params=[True, False], ids=["master_detail_true", "master_detail_false"])
def program(request, office, household_checker, individual_checker):
    from testutils.factories import CountryProgramFactory

    return CountryProgramFactory(
        country_office=office,
        household_checker=household_checker,
        individual_checker=individual_checker,
        household_columns="__str__\nid\nxx",
        individual_columns="__str__\nid\nxx",
        beneficiary_group__master_detail=request.param,
    )


@pytest.fixture
def household(program):
    from testutils.factories import CountryHouseholdFactory

    return CountryHouseholdFactory(batch__program=program, batch__country_office=program.country_office)


@pytest.fixture(params=["hh", "ind"])
def data(request: pytest.FixtureRequest, household: "CountryHousehold") -> tuple[io.BytesIO, "CountryHousehold", str]:
    def create_xlsx_buffer(header: list[str], rows: list[list[Any]]) -> io.BytesIO:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        for col, cell in enumerate(header):
            worksheet.write(0, col, cell)
        for row_idx, row in enumerate(rows, start=1):
            for col, cell in enumerate(row):
                worksheet.write(row_idx, col, cell)
        workbook.close()
        buffer.seek(0)
        return buffer

    def set_hh_fields(fields: list[str]) -> list[str]:
        return [household.id, household.version] + [household.flex_fields.get(field, None) for field in fields]

    def set_ind_fields(fields: list[str]) -> list[list[Any]]:
        return [
            [member.id, member.version] + [member.flex_fields.get(field, None) for field in fields]
            for member in household.members.all()
        ]

    target = request.param
    if target == "hh":
        buff = create_xlsx_buffer(
            stub.header_base + stub.header_add["hh"],
            [set_hh_fields(stub.header_add["hh"])],
        )
    elif target == "ind":
        buff = create_xlsx_buffer(
            stub.header_base + stub.header_add["ind"],
            set_ind_fields(stub.header_add["ind"]),
        )
    else:
        raise ValueError(f"Invalid target: {request.param}")

    return buff, household, target


@pytest.mark.parametrize(("field", "validator"), list(TYPES.items()))
def test_validator(field, validator):
    flex_field = FlexFieldFactory(definition__field_type=field, definition__attrs={"choices": [("a", "A")]})
    assert validator(flex_field)()


def test_create_bulk_update_template(household: "CountryHousehold", force_migrated_records):
    selected_fields = stub.header_base + stub.header_add["ind"]
    ret = create_bulk_update_template(
        household.members.all(),
        household.program,
        selected_fields,
    )
    workbook = openpyxl.load_workbook(io.BytesIO(ret.getvalue()))
    sheet = workbook.worksheets[0]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers == selected_fields


def test_bulk_update_export(
    app: "DjangoTestApp",
    force_migrated_records,
    household: "CountryHousehold",
    mock_storage,
) -> None:
    url = reverse("workspace:workspaces_countryindividual_changelist")
    selected_fields = stub.header_add["ind"]
    with mock.patch("country_workspace.workspaces.admin.cleaners.bulk_update.MEDIA_STORAGE", mock_storage):
        with select_office(app, household.country_office, household.program):
            res = app.get(url)
            form = res.forms["changelist-form"]
            form["action"] = "bulk_update_export"
            form.set("_selected_action", True, index=0)
            res = form.submit()

            form = res.forms["bulk-update-form"]
            for i in range(len(form.fields.get("fields"))):
                target: Checkbox = form.fields.get("fields")[i]
                if target._value in selected_fields:
                    target.checked = True
            res = form.submit("_export")

            assert res.status_code == 302
            job: AsyncJob = household.program.jobs.first()
            job.queue()


def test_bulk_update_import(
    app: "DjangoTestApp",
    force_migrated_records,
    household: "CountryHousehold",
    data: tuple[io.BytesIO, "CountryHousehold", str],
) -> None:
    buff, household, target = data
    url = reverse("workspace:workspaces_countryprogram_change", args=[household.program.pk])

    with select_office(app, household.country_office, household.program):
        res = app.get(url)
        res = res.click("Update Records")
        res.forms["bulk-update-form"]["description"] = f"Bulk update from {target}"
        res.forms["bulk-update-form"]["target"] = target
        res.forms["bulk-update-form"]["file"] = Upload(f"{target}.xlsx", buff.read())
        res = res.forms["bulk-update-form"].submit("_import")
        household.refresh_from_db()
        job: AsyncJob = household.program.jobs.first()

        assert res.status_code == 302
        assert job

        if target == "hh":
            admin1_v = household.flex_fields.get("admin_1")
            assert household.flex_fields.get("admin1") == admin1_v
        elif target == "ind":
            given_name = household.flex_fields.get("given_name")
            assert household.members.filter(flex_fields__given_name=given_name).exists()
